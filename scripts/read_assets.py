# -*- coding: utf-8 -*-
"""技術アセット棚卸しファイル(Excel/CSV)をパースしてassets.jsonを出力。

使い方:
  python read_assets.py --in 技術アセット棚卸.xlsx --out output/assets.json

対応フォーマット(棚卸し標準形):
  ヘッダー行: 名称/カテゴリ/成熟度/ステータス/最終更新/技術概要/[重要]xx|説明/[基本]xx/
              [機能]機能グループ > 機能名|コア・提供価値・制約条件・差別化要素
  - 1シートに「ヘッダー行+データ行」のブロックが複数並ぶ形式に対応(空行区切り)
  - CSVは1ブロックのみの想定
出力JSON(1アセット=1オブジェクト):
  {"asset_id","name","category","maturity","overview",
   "specs":[{"type":"重要/基本","name","value","desc"}],
   "functions":[{"group","name","core","value","constraint","differentiator"}]}
"""
import argparse, json, os, re, sys


def parse_block(header, row):
    asset = {"asset_id": None, "name": None, "category": None, "maturity": None,
             "overview": None, "specs": [], "functions": []}
    spec_map, func_map = {}, {}
    for h, v in zip(header, row):
        if h is None or v is None or str(v).strip() in ("", "nan", "NaT"):
            continue
        h, v = str(h).strip(), str(v).strip()
        if h == "名称":
            asset["name"] = v; asset["asset_id"] = v
        elif h == "カテゴリ":
            asset["category"] = v
        elif h == "成熟度":
            asset["maturity"] = v
        elif h == "技術概要":
            asset["overview"] = v
        elif h.startswith("[重要]") or h.startswith("[基本]"):
            typ = "重要" if h.startswith("[重要]") else "基本"
            body_ = h[4:]
            if "｜説明" in body_ or "|説明" in body_:
                key = re.split(r"[｜|]", body_)[0]
                spec_map.setdefault((typ, key), {})["desc"] = v
            else:
                spec_map.setdefault((typ, body_), {})["value"] = v
        elif h.startswith("[機能]"):
            body_ = h[4:]
            parts = re.split(r"[｜|]", body_)
            if len(parts) != 2:
                continue
            path, attr = parts[0].strip(), parts[1].strip()
            if ">" in path:
                grp, fname = [x.strip() for x in path.split(">", 1)]
            else:
                grp, fname = "", path
            key = (grp, fname)
            func_map.setdefault(key, {})
            attr_key = {"コア": "core", "提供価値": "value", "制約条件": "constraint",
                        "差別化要素": "differentiator"}.get(attr)
            if attr_key:
                func_map[key][attr_key] = v
    for (typ, name), d in spec_map.items():
        asset["specs"].append({"type": typ, "name": name,
                               "value": d.get("value"), "desc": d.get("desc")})
    for (grp, fname), d in func_map.items():
        asset["functions"].append({"group": grp, "name": fname, **{k: d.get(k) for k in
                                   ("core", "value", "constraint", "differentiator")}})
    return asset if asset["name"] else None


def parse_transcribed_sheet(rows):
    """build_startup_db.pyが生成した⑤アセット情報シート(転記済み)を逆パースする。
    元の棚卸しファイルが手元になく、マッチング用xlsxの転記シートしかない場合に使う。
    先頭列が「■ 名称」で始まるブロックの繰り返し(カテゴリ/成熟度/技術概要 → スペック表 → 機能表 → WANT/NG)。
    """
    assets = []
    asset = None
    section = None  # "spec" | "func" | None
    for row in rows:
        cells = ["" if c is None else str(c).strip() for c in row]
        c0 = cells[0] if cells else ""
        if not any(cells):
            continue
        if c0.startswith("■ "):
            if asset and asset["name"]:
                assets.append(asset)
            asset = {"asset_id": None, "name": c0[2:].strip(), "category": None,
                      "maturity": None, "overview": None, "specs": [], "functions": [],
                      "expansion_wants": [], "ng_conditions": []}
            asset["asset_id"] = asset["name"]
            section = None
            continue
        if asset is None:
            continue
        if c0 == "カテゴリ":
            asset["category"] = cells[1] if len(cells) > 1 else None
        elif c0 == "成熟度":
            asset["maturity"] = cells[1] if len(cells) > 1 else None
        elif c0 == "技術概要":
            asset["overview"] = cells[1] if len(cells) > 1 else None
        elif c0 == "スペック項目":
            section = "spec"
        elif c0.startswith("機能(機能グループ"):
            section = "func"
        elif c0 == "拡張可能性(WANT)":
            v = cells[1] if len(cells) > 1 else ""
            if v:
                asset["expansion_wants"].append(v)
        elif c0 == "NG条件":
            v = cells[1] if len(cells) > 1 else ""
            if v:
                asset["ng_conditions"].append(v)
        elif section == "spec" and c0.startswith("["):
            m = re.match(r"\[(.+?)\]\s*(.*)", c0)
            typ, name = (m.group(1), m.group(2)) if m else ("基本", c0)
            asset["specs"].append({"type": typ, "name": name,
                                   "value": cells[1] if len(cells) > 1 else None,
                                   "desc": cells[2] if len(cells) > 2 else None})
        elif section == "func" and c0:
            grp, _, fname = c0.partition(" > ")
            if not fname:
                grp, fname = "", c0
            asset["functions"].append({
                "group": grp, "name": fname,
                "core": cells[1] if len(cells) > 1 and cells[1] else None,
                "value": cells[2] if len(cells) > 2 and cells[2] else None,
                "constraint": cells[3] if len(cells) > 3 and cells[3] else None,
                "differentiator": cells[4] if len(cells) > 4 and cells[4] else None,
            })
    if asset and asset["name"]:
        assets.append(asset)
    return assets


def parse_form_sheet(rows, sheet_name):
    """記入ページ形式(1.技術概要/2.基本スペック/3.機能分解/4.拡張可能性/5.NG条件)をパース。"""
    asset = {"asset_id": None, "name": None, "category": None, "maturity": None,
             "overview": None, "specs": [], "functions": [],
             "one_liner": None, "main_use": None, "strengths": None,
             "expansion_wants": [], "ng_conditions": []}
    section = None
    func = None
    grid = [["" if c is None else str(c).strip() for c in r] for r in rows]
    for r in grid:
        vals = [c for c in r if c]
        if not vals:
            continue
        joined = vals[0]
        if joined.startswith("1.") and "技術概要" in joined:
            section = "overview"; continue
        if joined.startswith("2.") and "スペック" in joined:
            section = "spec"; continue
        if joined.startswith("3.") and "機能分解" in joined:
            section = "func"; continue
        if (joined.startswith("4") or joined.startswith("４")) and "拡張" in joined:
            section = "want"; continue
        if (joined.startswith("5") or joined.startswith("５")) and "NG" in joined:
            section = "ng"; continue
        key, val = vals[0], (vals[1] if len(vals) > 1 else None)
        if key in ("項目", "内容") and not val:
            continue
        if section == "overview":
            if key == "技術名":
                asset["name"] = val; asset["asset_id"] = val
            elif key == "一言説明":
                asset["one_liner"] = val
            elif key == "主用途":
                asset["main_use"] = val
            elif key == "強み":
                asset["strengths"] = val
        elif section == "spec":
            if val and key != "項目" and not key.startswith("　"):
                asset["specs"].append({"type": "基本", "name": key, "value": val, "desc": None})
                if "実用化" in key or "成熟" in key:
                    asset["maturity"] = val
        elif section == "func":
            if key[0] in "①②③④⑤⑥⑦⑧⑨" or (key[0].isdigit() and len(key) > 2):
                func = {"group": "", "name": key.lstrip("①②③④⑤⑥⑦⑧⑨0123456789.、 "),
                        "core": None, "value": None, "constraint": None, "differentiator": None}
                asset["functions"].append(func)
            elif func is not None and val:
                m = {"提供価値": "value", "制約条件": "constraint", "差別化要素": "differentiator",
                     "コア": "core"}.get(key)
                if m:
                    func[m] = val
        elif section == "want" and key != "内容":
            asset["expansion_wants"].append(key)
        elif section == "ng" and key != "内容":
            asset["ng_conditions"].append(key)
    if asset["name"]:
        parts = [x for x in (asset["one_liner"], ("主用途: " + asset["main_use"]) if asset["main_use"] else None,
                             ("強み: " + asset["strengths"]) if asset["strengths"] else None) if x]
        asset["overview"] = "。".join(parts) or None
        return asset
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True)
    ap.add_argument("--out", default="output/assets.json")
    a = ap.parse_args()

    rows_by_sheet = {}
    if a.inp.lower().endswith(".csv"):
        import csv
        with open(a.inp, encoding="utf-8-sig", newline="") as f:
            rows_by_sheet["csv"] = list(csv.reader(f))
    else:
        from openpyxl import load_workbook
        wb = load_workbook(a.inp, data_only=True)
        for ws in wb.worksheets:
            rows_by_sheet[ws.title] = [[c for c in row] for row in ws.iter_rows(values_only=True)]

    assets, seen = [], set()
    for sheet, rows in rows_by_sheet.items():
        if "記載例" in sheet:  # テンプレートの記入例シートはスキップ
            continue
        a1 = str(rows[0][0]) if rows and rows[0] and rows[0][0] else ""
        # 転記済み形式の検出(build_startup_db.pyが生成した⑤アセット情報シート。A1が「アセット情報(N件)」)
        if "アセット情報" in a1:
            for asset in parse_transcribed_sheet(rows):
                if asset["name"] not in seen:
                    seen.add(asset["name"])
                    assets.append(asset)
            continue
        # 記入ページ形式の検出(セルに「技術概要」セクション見出しがある)
        flat = " ".join(str(c) for r in rows[:15] for c in r if c)
        if "技術概要" in flat and "技術名" in " ".join(str(c) for r in rows for c in r if c):
            asset = parse_form_sheet(rows, sheet)
            if asset and asset["name"] not in seen:
                seen.add(asset["name"])
                assets.append(asset)
            continue
        # 棚卸し標準形(ヘッダー行「名称」)
        header = None
        for row in rows:
            cells = ["" if c is None else str(c).strip() for c in row]
            if not any(cells):
                continue
            if cells[0] == "名称":  # 新しいブロックのヘッダー
                header = row
                continue
            if header is not None:
                asset = parse_block(header, row)
                if asset and asset["name"] not in seen:
                    seen.add(asset["name"])
                    assets.append(asset)

    if not assets:
        sys.exit("アセットを1件も検出できませんでした。ヘッダー行(先頭列「名称」)があるか確認してください")
    os.makedirs(os.path.dirname(a.out) or ".", exist_ok=True)
    json.dump(assets, open(a.out, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print(f"{len(assets)}アセットを {a.out} に保存:")
    for x in assets:
        print(f"  - {x['name']} | {x['category']} | {x['maturity']} | 機能{len(x['functions'])}件 スペック{len(x['specs'])}件")


if __name__ == "__main__":
    main()
