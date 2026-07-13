# -*- coding: utf-8 -*-
"""スタートアップDB(40項目)+マッチングDB(キー4列+本編30列+要約ブロック13列=47列)のExcel生成。

使い方:
  python build_startup_db.py --in output/startups.jsonl --out スタートアップDB.xlsx
  python build_startup_db.py --in output/startups.jsonl --matching output/matching.jsonl --out スタートアップDB.xlsx
  python build_startup_db.py --template   # 空テンプレートを templates/ に生成

シート構成: ①サマリ / ②項目定義 / ③スタートアップDB / ④マッチングDB(--matching時またはテンプレート)
マッチングDBのスコアは0-10の数値のみ("8/10"等の文字列はExcelが日付化するため禁止)。
自己評価ランクはscore_totalから自動導出(A>=8.0 / B>=6.0 / C<6.0)。
"""
import argparse, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Meiryo UI"
NAVY = "1F3864"; GOLD = "7F6000"; LBLUE = "DCE6F1"; YEL = "FFF2CC"; LGRAY = "F2F2F2"
GREEN = "E2EFDA"; ORANGE = "FCE4D6"
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

INDUSTRY_CATS = [
    "AI・データ基盤", "エンタープライズSaaS・業務DX", "フィンテック・保険", "ヘルスケア・医療機器",
    "バイオ・創薬", "モビリティ・自動車", "宇宙・航空・防衛", "エネルギー・クリーンテック",
    "素材・化学", "半導体・エレクトロニクス", "ロボティクス・ドローン", "製造DX・産業機器",
    "物流・サプライチェーン", "農業・食品", "建設・不動産・インフラ", "小売・EC・消費者サービス",
    "HR・教育", "セキュリティ・通信", "エンタメ・メディア・ゲーム", "その他",
]
MFG_CATS = ["AI・データ(シミュレーション/CAE)", "センシング", "検査・品質", "予知保全",
            "ロボティクス・FA", "エッジAI", "材料・MI", "熱・パワエレ"]
STAGE_VOCAB = ["シード", "シリーズA", "シリーズB", "シリーズC", "シリーズD以降", "グラント・デットのみ", "不明"]

# ---------- ③スタートアップDB(No+40項目) ----------
DB_COLS = [  # (header, width, json_field)
    ("No", 5, None), ("スタートアップID", 18, "startup_id"),
    ("業界分類", 20, "industry_category"), ("製造技術分類", 16, "mfg_tech_category"),
    ("企業名", 22, "name"), ("企業名(日)", 20, "name_ja"), ("国", 10, "country"), ("都市", 13, "hq_city"),
    ("設立年", 8, "founded_year"), ("Webサイト", 24, "website"), ("従業員(概算)", 12, "employees"),
    ("ステージ(原文)", 16, "stage"), ("ステージ(統一)", 13, "stage_normalized"),
    ("累計調達額", 18, "total_funding"), ("調達額USD(百万)", 12, "funding_usd"),
    ("直近ラウンド", 30, "last_round"), ("直近調達日", 11, "last_round_date"),
    ("主要投資家", 28, "key_investors"),
    ("事業概要", 42, "business_summary"), ("主要プロダクト", 20, "product_name"),
    ("プロダクト概要", 42, "product_summary"), ("提供形態", 18, "delivery_model"),
    ("収益モデル", 20, "revenue_model"), ("想定顧客", 30, "target_customer"),
    ("対象業界", 20, "target_industries"), ("主要顧客・パートナー", 24, "notable_customers_partners"),
    ("課題", 40, "customer_pain"),
    ("市場規模", 16, "market_size"), ("市場規模詳細", 34, "market_size_detail"),
    ("CAGR", 12, "market_cagr"), ("CAGR詳細", 30, "market_cagr_detail"),
    ("技術キーワード", 26, "tech_keywords"), ("コア技術", 42, "core_tech_description"),
    ("差別化要素", 38, "differentiator"), ("成熟度", 16, "maturity"),
    ("定義適合", 8, "fit"), ("適合注記", 22, "fit_note"), ("主要リスク(事実)", 30, "key_risks"),
    ("出典URL", 28, "source_urls"), ("情報取得日", 11, "info_date"), ("取得経路", 13, "collect_route"),
]

# ---------- ④マッチングDB(キー4+本編30+要約13=47列) ----------
M_COLS = [  # (header, width, json_field or callable-tag)  _db_*はスタートアップDBからの自動転記
    ("スタートアップID", 18, "startup_id"), ("企業名", 22, "startup_name"),
    ("事業概要", 40, "_db_business_summary"), ("主要プロダクト", 20, "_db_product_name"),
    ("プロダクト概要", 40, "_db_product_summary"), ("コア技術", 40, "_db_core_tech"),
    ("アセットID(名)", 26, "asset_id"), ("協業タイプ", 16, "coop_type"),
    ("No", 5, "_no"), ("タイトル", 30, "title"), ("タグ", 24, "tags"), ("概要", 44, "summary"),
    ("市場規模(一覧)", 14, "market_size_list"), ("CAGR(一覧)", 11, "cagr_list"),
    ("総合スコア(一覧)", 10, "_total_list"), ("自己評価ランク", 9, "_rank"),
    ("想定顧客", 30, "target_customer"), ("課題", 38, "pain"), ("解決方法", 44, "solution"),
    ("収益モデル", 28, "revenue_model"), ("アイデアの強み", 38, "idea_strength"),
    ("使用するコア技術", 42, "core_tech_used"),
    ("差別化スコア", 8, "score_differentiation"), ("技術実現性スコア", 8, "score_tech_feasibility"),
    ("アセット適合性スコア", 8, "score_asset_fit"), ("業界優位性スコア", 8, "score_industry_advantage"),
    ("業界課題フィットスコア", 8, "score_industry_pain_fit"),
    ("未使用理由・採用障壁", 34, "unused_reason_barriers"), ("コア機能の改善点", 34, "core_improvement"),
    ("新規性スコア", 8, "score_novelty"), ("ミッション整合性スコア", 8, "score_mission_fit"),
    ("市場規模詳細", 30, "market_size_detail"), ("市場規模スコア", 8, "score_market_size"),
    ("CAGR詳細", 28, "cagr_detail"), ("CAGRスコア", 8, "score_cagr"),
    ("総合評価スコア", 10, "score_total"), ("エージェントの思考根拠", 46, "agent_rationale"),
    ("主要リスク", 38, "key_risks"),
    # ---- 要約ブロック(13列): 一覧・他システム連携用の凝縮表記 ----
    ("要約_コンセプト", 26, "sum_concept"), ("要約_想定市場", 22, "sum_market"),
    ("要約_顧客", 24, "sum_customer"), ("要約_課題", 30, "sum_pain"),
    ("要約_解決方法", 34, "sum_solution"), ("要約_市場規模", 18, "sum_market_size"),
    ("要約_CAGR", 12, "sum_cagr"), ("要約_独自性", 30, "sum_uniqueness"),
    ("要約_ミッション整合", 28, "sum_mission"),
    ("要約_独自性スコア", 8, "score_sum_uniqueness"),
    ("要約_ミッション整合スコア", 8, "score_sum_mission"),
    ("要約_市場規模スコア", 8, "score_sum_market_size"),
    ("要約_成長性スコア", 8, "score_sum_growth"),
]
SCORE_FIELDS = {"score_differentiation", "score_tech_feasibility", "score_asset_fit",
                "score_industry_advantage", "score_industry_pain_fit", "score_novelty",
                "score_mission_fit", "score_market_size", "score_cagr", "score_total",
                "_total_list", "score_sum_uniqueness", "score_sum_mission",
                "score_sum_market_size", "score_sum_growth"}

DB_ITEM_DEFS = [
    ("startup_id", "一意ID(国内=法人番号、無ければjp-社名/海外=ドメイン)", "マッチングDBとの結合キー"),
    ("業界分類", "20分類から1つ(包括的タクソノミー)", "突合①: 粗い絞り込み"),
    ("製造技術分類", "8分類、製造業関連の場合のみ", "突合①: 製造系アセットの絞り込み"),
    ("企業名/企業名(日)", "英名/日本語名(国内のみ)", "(識別)"),
    ("国/都市", "国は日本語表記", "実現性評価: 地理"),
    ("設立年/Webサイト/従業員", "事実のみ。従業員は「約N名(推定)」可", "定義判定・体制規模"),
    ("ステージ(原文/統一)", "統一語彙: シード/シリーズA/B/C/D以降/グラント・デットのみ/不明", "協業タイプの当たり付け・集計"),
    ("累計調達額/調達額USD(百万)", "原文表記+USD換算数値(合算・換算は明記)", "ソート・優先度付け"),
    ("直近ラウンド/直近調達日", "「YYYY年M月・金額・ラウンド名」+ISO日付", "鮮度・勢い"),
    ("主要投資家", "リード優先", "競合CVCの確認"),
    ("事業概要/主要プロダクト/プロダクト概要", "日本語1-2文", "◎マッチング対象の実体"),
    ("提供形態/収益モデル", "SaaS等/サブスク・部材販売等(収益の取り方)", "協業スキーム・商流設計"),
    ("想定顧客/対象業界/主要顧客", "具体的な顧客像/業界リスト/公開情報のみ", "顧客候補型の判定・実績確認"),
    ("課題", "その企業が解く顧客課題", "◎アセットの「提供価値」と対照"),
    ("市場規模/詳細・CAGR/詳細", "出典レポートが確認できた場合のみ記入(null許容)", "市場規模・CAGRスコアの根拠"),
    ("技術キーワード", "統制語彙+自由タグ、3-6個", "◎技術突合キー"),
    ("コア技術", "方式・原理・定量性能のいずれか必須", "◎「使用するコア技術」の材料"),
    ("差別化要素", "技術・事業の強み(受賞歴等は不可)", "競合/補完の判定"),
    ("成熟度", "PoC/製品提供中/量産採用あり", "実現性評価"),
    ("定義適合/適合注記", "○/△。△は定義6基準上の理由のみ", "足切り"),
    ("主要リスク(事実)", "買収・訴訟・依存等の事実ベースのみ", "リスク評価の材料"),
    ("出典URL/情報取得日/取得経路", "検証可能性と再現性の記録", "裏取り・鮮度管理"),
]
M_ITEM_DEFS = [
    ("キー4列(スタートアップID/企業名/アセットID/協業タイプ)", "両DBへの結合キー。協業タイプ=技術提供・ライセンス/共同開発/調達・ツール導入/顧客候補/出資"),
    ("転記4列(事業概要/主要プロダクト/プロダクト概要/コア技術)", "③スタートアップDBからstartup_idで自動転記(Excel生成時)。マッチした相手の実体をこのシート内で確認できるようにするための参照列。手で書かない"),
    ("⑤アセット情報シート", "マッチングに使用したアセットの全情報(名称/成熟度/概要/スペック/機能分解/拡張可能性/NG条件)を転記して保存。build_startup_db.py --assets で生成"),
    ("No/タイトル/タグ/概要", "協業アイデアの識別・要約(概要=アセットの何を相手の何と組み合わせ何を生むか)"),
    ("市場規模(一覧)/CAGR(一覧)/市場規模詳細/CAGR詳細", "スタートアップDBから転記(協業固有の市場ならその値・出典必須)"),
    ("総合スコア(一覧)/総合評価スコア", "同値(一覧表示用と詳細用)。スコア10種の加重平均・0-10・小数1桁"),
    ("自己評価ランク", "総合評価スコアから自動導出: A=8.0以上/B=6.0〜7.9/C=6.0未満"),
    ("想定顧客/課題/解決方法/収益モデル/アイデアの強み", "協業アイデアとしての記述(主語=アセット×スタートアップ)"),
    ("使用するコア技術", "「アセット側: …/スタートアップ側: …」の形式で両方明記"),
    ("スコア9種(各0-10の数値)", "差別化(補完性含む: 相手と重複=競合=低)/技術実現性/アセット適合性/業界優位性/業界課題フィット/新規性/ミッション整合性/市場規模/CAGR。※「8/10」等の文字列は日付化するため禁止"),
    ("未使用理由・採用障壁/コア機能の改善点", "アセットの制約条件・協業の障壁/協業で埋めるべきギャップ"),
    ("エージェントの思考根拠", "スコアの根拠。事実(両者の技術・実績・出典)のみ引用"),
    ("主要リスク", "協業アイデアのリスク(評価)"),
    ("要約ブロック13列(要約_コンセプト〜要約_成長性スコア)", "一覧・他システム連携用の凝縮表記。コンセプト/想定市場/顧客/課題/解決方法/市場規模/CAGR/独自性/ミッション整合(各1文以内)+スコア4種(独自性/ミッション整合/市場規模/成長性、0-10数値)。本編列の内容を要約したもので矛盾させないこと"),
]


def hdr(cell, fill=NAVY):
    cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", start_color=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border


def body(cell, fill=None, center=False, bold=False):
    cell.font = Font(name=FONT, size=10, bold=bold)
    cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal=("center" if center else None))
    cell.border = border
    if fill:
        cell.fill = PatternFill("solid", start_color=fill)


def title(ws, text, sub=None):
    c = ws.cell(1, 1, text)
    c.font = Font(name=FONT, bold=True, size=14, color=NAVY)
    if sub:
        s = ws.cell(2, 1, sub)
        s.font = Font(name=FONT, size=9, color="595959")


def rank_of(total):
    if total is None:
        return None
    return "A" if total >= 8.0 else ("B" if total >= 6.0 else "C")


def build(recs, out_path, matching=None, matching_note=None, assets=None):
    wb = Workbook()
    # マッチング行への転記用: startup_id/企業名 → DBレコード
    db_by_key = {}
    for d in recs:
        for k in (d.get("startup_id"), d.get("name"), d.get("name_ja")):
            if k:
                db_by_key[str(k).lower()] = d

    # ---- ③ スタートアップDB ----
    db = wb.active
    db.title = "③スタートアップDB"
    title(db, f"スタートアップDB({len(recs)}社)",
          "アセット非依存のマスタDB(事実のみ)。評価・マッチングは④マッチングDBで管理。")
    r = 4
    for i, (h, w, _) in enumerate(DB_COLS, 1):
        hdr(db.cell(r, i, h))
        db.column_dimensions[get_column_letter(i)].width = w
    for no, d in enumerate(recs, 1):
        r += 1
        jp = d.get("country") == "日本"
        for i, (h, w, field) in enumerate(DB_COLS, 1):
            if field is None:
                v = no
            else:
                v = d.get(field)
                if isinstance(v, list):
                    v = ("\n" if field == "source_urls" else "、").join(str(x) for x in v)
                if v is None:
                    v = ""
            c = db.cell(r, i, v)
            body(c, center=(field is None))
            if jp and field == "country":
                c.fill = PatternFill("solid", start_color=YEL)
            if field == "fit_note" and d.get("fit") == "△":
                c.fill = PatternFill("solid", start_color=YEL)
    db.freeze_panes = "E5"
    db.auto_filter.ref = f"A4:{get_column_letter(len(DB_COLS))}{max(r, 5)}"
    db_last = max(r, 5)

    # ---- ④ マッチングDB ----
    m_last = 5
    ms = wb.create_sheet("④マッチングDB")
    title(ms, f"マッチングDB({len(matching or [])}件)",
          "アセット×スタートアップの協業アイデア評価(アイデア評価DB互換フォーマット)。スコアは0-10の数値。ランクは総合評価スコアから自動導出(A>=8.0/B>=6.0/C<6.0)。")
    r = 4
    for i, (h, w, _) in enumerate(M_COLS, 1):
        hdr(ms.cell(r, i, h), fill=GOLD)
        ms.column_dimensions[get_column_letter(i)].width = w
    rank_fill = {"A": GREEN, "B": LBLUE, "C": ORANGE}
    for no, d in enumerate(matching or [], 1):
        r += 1
        total = d.get("score_total")
        rank = rank_of(total)
        src = (db_by_key.get(str(d.get("startup_id") or "").lower())
               or db_by_key.get(str(d.get("startup_name") or "").lower()) or {})
        for i, (h, w, field) in enumerate(M_COLS, 1):
            if field == "_no":
                v = no
            elif field == "_total_list":
                v = total if total is not None else ""
            elif field == "_rank":
                v = rank or ""
            elif field == "_db_business_summary":
                v = src.get("business_summary") or ""
            elif field == "_db_product_name":
                v = src.get("product_name") or ""
            elif field == "_db_product_summary":
                v = src.get("product_summary") or ""
            elif field == "_db_core_tech":
                v = src.get("core_tech_description") or ""
            else:
                v = d.get(field)
                if isinstance(v, list):
                    v = "; ".join(str(x) for x in v)
                if v is None:
                    v = ""
            c = ms.cell(r, i, v)
            body(c, center=(field in SCORE_FIELDS or field in ("_no", "_rank")))
            if field == "_rank" and rank:
                c.fill = PatternFill("solid", start_color=rank_fill[rank])
                c.font = Font(name=FONT, size=10, bold=True)
    ms.freeze_panes = "E5"
    ms.auto_filter.ref = f"A4:{get_column_letter(len(M_COLS))}{max(r, 5)}"
    m_last = max(r, 5)

    # ---- ① サマリ ----
    sm = wb.create_sheet("①サマリ", 0)
    title(sm, "サマリ", "スタートアップ定義(収載基準)+③④の自動集計")
    r = 4
    c = sm.cell(r, 1, "A. スタートアップ定義(収載6基準)"); c.font = Font(name=FONT, bold=True, size=12, color=NAVY)
    defs = [
        (1, "未上場", "上場企業・連結子会社・買収済みは除外(カーブアウトは可)"),
        (2, "設立年数", "10年以内(ディープテックは15年以内まで許容)"),
        (3, "外部資金調達", "エクイティ調達実績あり(国内は必須。助成金・融資のみは△)"),
        (4, "成長志向", "スケール可能な製品・サービス(受託専業は除外)"),
        (5, "規模", "従業員おおむね500名未満(目安)"),
        (6, "独立性", "大企業グループ中核子会社・JVは原則除外"),
    ]
    r += 1
    for i, h in enumerate(["#", "基準", "内容"], 1):
        hdr(sm.cell(r, i, h))
    for row in defs:
        r += 1
        for i, v in enumerate(row, 1):
            body(sm.cell(r, i, v), center=(i == 1))
        sm.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)

    r += 2
    c = sm.cell(r, 1, "B. スタートアップDB集計(業界分類×国内外)"); c.font = Font(name=FONT, bold=True, size=12, color=NAVY)
    r += 1
    for i, h in enumerate(["業界分類", "国内", "海外", "合計"], 1):
        hdr(sm.cell(r, i, h))
    r0 = r + 1
    for cat in INDUSTRY_CATS:
        r += 1
        body(sm.cell(r, 1, cat), bold=True)
        sm.cell(r, 2, f"=COUNTIFS('③スタートアップDB'!C5:C{db_last},A{r},'③スタートアップDB'!G5:G{db_last},\"日本\")")
        sm.cell(r, 3, f"=COUNTIFS('③スタートアップDB'!C5:C{db_last},A{r})-B{r}")
        sm.cell(r, 4, f"=B{r}+C{r}")
        for i in (2, 3, 4):
            body(sm.cell(r, i), center=True)
    r += 1
    body(sm.cell(r, 1, "合計"), bold=True, fill=LBLUE)
    for col, L in ((2, "B"), (3, "C"), (4, "D")):
        sm.cell(r, col, f"=SUM({L}{r0}:{L}{r-1})")
        body(sm.cell(r, col), bold=True, fill=LBLUE, center=True)
    total_row = r
    r += 1
    body(sm.cell(r, 1, "カバー分類数"), bold=True, fill=YEL)
    sm.cell(r, 2, f"=SUMPRODUCT((D{r0}:D{total_row-1}>0)*1)")
    body(sm.cell(r, 2), bold=True, fill=YEL, center=True)
    body(sm.cell(r, 3, f"/ {len(INDUSTRY_CATS)}分類"), fill=YEL)

    r += 2
    c = sm.cell(r, 1, "C. 製造技術分類(追加タグ)の集計"); c.font = Font(name=FONT, bold=True, size=12, color=NAVY)
    r += 1
    for i, h in enumerate(["製造技術分類", "件数"], 1):
        hdr(sm.cell(r, i, h))
    m0 = r + 1
    for cat in MFG_CATS:
        r += 1
        body(sm.cell(r, 1, cat), bold=True)
        sm.cell(r, 2, f"=COUNTIF('③スタートアップDB'!D5:D{db_last},A{r})")
        body(sm.cell(r, 2), center=True)
    r += 1
    body(sm.cell(r, 1, "製造業関連 計"), bold=True, fill=LBLUE)
    sm.cell(r, 2, f"=SUM(B{m0}:B{r-1})")
    body(sm.cell(r, 2), bold=True, fill=LBLUE, center=True)

    r += 2
    c = sm.cell(r, 1, "D. マッチングDB集計"); c.font = Font(name=FONT, bold=True, size=12, color=GOLD)
    r += 1
    for i, h in enumerate(["項目", "値"], 1):
        hdr(sm.cell(r, i, h), fill=GOLD)
    r += 1
    body(sm.cell(r, 1, "マッチング件数"), bold=True)
    sm.cell(r, 2, f"=COUNTA('④マッチングDB'!A5:A{max(m_last,5)})")
    body(sm.cell(r, 2), center=True)
    rank_col = get_column_letter([f for _, _, f in M_COLS].index("_rank") + 1)  # 列位置はM_COLSから導出(列追加でズレない)
    for rk in ("A", "B", "C"):
        r += 1
        body(sm.cell(r, 1, f"ランク{rk}"), bold=True, fill={"A": GREEN, "B": LBLUE, "C": ORANGE}[rk])
        sm.cell(r, 2, f"=COUNTIF('④マッチングDB'!{rank_col}5:{rank_col}{max(m_last,5)},\"{rk}\")")
        body(sm.cell(r, 2), center=True)
    r += 1
    body(sm.cell(r, 1, "選定方式(既定)"), bold=True, fill=YEL)
    body(sm.cell(r, 2, "マッチング件数はアセットごとの上位N選抜(既定: 各5〜10ペア)。選外の企業は『除外』ではなく『未採点』であり、スコアによる足切りは行っていない。fit=△の企業も自動除外しない(リスクは主要リスク欄に記載)。"))
    sm.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    if matching_note:
        r += 1
        body(sm.cell(r, 1, "本ファイルの選定条件"), bold=True, fill=YEL)
        body(sm.cell(r, 2, matching_note))
        sm.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    for i, w in enumerate([30, 12, 10, 10, 26], 1):
        sm.column_dimensions[get_column_letter(i)].width = w

    # ---- ② 項目定義 ----
    it = wb.create_sheet("②項目定義", 1)
    title(it, "項目定義", "上段=スタートアップDB(40項目・事実のみ)/下段=マッチングDB(キー4列+本編30列+要約13列・評価)。詳細ルールはスキルSKILL.md。")
    r = 4
    c = it.cell(r, 1, "A. スタートアップDB(40項目)"); c.font = Font(name=FONT, bold=True, size=12, color=NAVY)
    r += 1
    for i, h in enumerate(["フィールド", "内容・記入ルール", "マッチングでの使われ方"], 1):
        hdr(it.cell(r, i, h))
    for row in DB_ITEM_DEFS:
        r += 1
        for i, v in enumerate(row, 1):
            fill = LBLUE if i == 1 else (YEL if i == 3 and v.startswith("◎") else None)
            body(it.cell(r, i, v), fill=fill)
    r += 2
    c = it.cell(r, 1, "B. マッチングDB(キー4列+本編30列+要約ブロック13列)"); c.font = Font(name=FONT, bold=True, size=12, color=GOLD)
    r += 1
    for i, h in enumerate(["列(グループ)", "内容・記入ルール"], 1):
        hdr(it.cell(r, i, h), fill=GOLD)
    for row in M_ITEM_DEFS:
        r += 1
        body(it.cell(r, 1, row[0]), fill=YEL)
        body(it.cell(r, 2, row[1]))
        it.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    for i, w in enumerate([44, 64, 44], 1):
        it.column_dimensions[get_column_letter(i)].width = w

    # ---- ⑤ アセット情報(マッチングに使用したアセットの転記) ----
    if assets:
        aw = wb.create_sheet("⑤アセット情報")
        title(aw, f"アセット情報({len(assets)}件)",
              "マッチングDB生成に使用した技術アセットの転記(read_assets.pyのパース結果)。マッチング根拠の参照用。")
        r = 4
        for x in assets:
            c = aw.cell(r, 1, f"■ {x.get('name')}")
            c.font = Font(name=FONT, bold=True, size=12, color=GOLD)
            r += 1
            meta = [("カテゴリ", x.get("category")), ("成熟度", x.get("maturity")),
                    ("技術概要", x.get("overview"))]
            for k, v in meta:
                if v:
                    body(aw.cell(r, 1, k), bold=True, fill=YEL)
                    body(aw.cell(r, 2, v))
                    aw.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
                    r += 1
            if x.get("specs"):
                for i, h in enumerate(["スペック項目", "値", "説明"], 1):
                    hdr(aw.cell(r, i, h), fill=GOLD)
                r += 1
                for s in x["specs"]:
                    body(aw.cell(r, 1, f"[{s.get('type','')}] {s.get('name','')}"))
                    body(aw.cell(r, 2, s.get("value") or ""))
                    body(aw.cell(r, 3, s.get("desc") or ""))
                    aw.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
                    r += 1
            if x.get("functions"):
                for i, h in enumerate(["機能(機能グループ>機能)", "コア", "提供価値", "制約条件", "差別化要素"], 1):
                    hdr(aw.cell(r, i, h), fill=GOLD)
                r += 1
                for f_ in x["functions"]:
                    nm = (f_.get("group") + " > " if f_.get("group") else "") + (f_.get("name") or "")
                    for i, v in enumerate([nm, f_.get("core"), f_.get("value"),
                                           f_.get("constraint"), f_.get("differentiator")], 1):
                        body(aw.cell(r, i, v or ""))
                    r += 1
            for label, key in (("拡張可能性(WANT)", "expansion_wants"), ("NG条件", "ng_conditions")):
                vals = x.get(key) or []
                if vals:
                    body(aw.cell(r, 1, label), bold=True, fill=YEL)
                    body(aw.cell(r, 2, " / ".join(vals)))
                    aw.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
                    r += 1
            r += 1  # アセット間の空行
        for i, w in enumerate([34, 30, 34, 30, 34, 20], 1):
            aw.column_dimensions[get_column_letter(i)].width = w

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", help="スタートアップDBのJSONL")
    ap.add_argument("--matching", help="マッチングDBのJSONL(任意)")
    ap.add_argument("--matching-note", dest="matching_note", help="サマリに記載する本実行の選定条件(任意)")
    ap.add_argument("--assets", help="read_assets.pyのassets.json(任意)。⑤アセット情報シートに転記")
    ap.add_argument("--out", help="出力xlsxパス")
    ap.add_argument("--template", action="store_true")
    a = ap.parse_args()

    def load(p):
        rows = []
        with open(p, encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    rows.append(json.loads(line))
        return rows

    if a.template:
        out = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "templates", "スタートアップDB_テンプレート.xlsx"))
        build([], out, matching=[])
        print(f"テンプレート生成: {out}")
        return
    if not a.inp or not a.out:
        raise SystemExit("--in と --out を指定するか、--template を使ってください")
    recs = load(a.inp)
    matching = load(a.matching) if a.matching else []
    assets = None
    if a.assets:
        with open(a.assets, encoding="utf-8") as f:
            assets = json.load(f)
    build(recs, a.out, matching=matching, matching_note=a.matching_note, assets=assets)
    print(f"保存: {a.out} (スタートアップDB {len(recs)}社 / マッチングDB {len(matching)}件"
          + (f" / アセット{len(assets)}件" if assets else "") + ")")


if __name__ == "__main__":
    main()
