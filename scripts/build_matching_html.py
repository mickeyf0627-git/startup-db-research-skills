#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""build_matching_html.py — マッチングExcel → 閲覧用HTMLレポート

能力2で生成した Excel(④マッチングDB / ⑤アセット情報 / ③スタートアップDB)を読み、
1ファイル完結の静的HTMLレポートを生成する。テンプレート(CSS/JS)は同ディレクトリの
matching_html_style.css / matching_html_app.js を読み込む。

構造:
- 上部: 対象技術アセット一覧(クリックで絞り込み) + フィルタバー
- カード: 「技術アセット列 | スタートアップ列」の2段組(下端に「本協業で使う技術」を揃える)
          → 協業アイデア(課題/解決策/強み/市場規模) → 折りたたみ(詳細スコア・想定顧客・
             収益モデル・評価根拠・採用障壁・改善点・リスク・市場データ詳細)

使い方(最小): 入力Excelだけ渡せば、同じ場所に「マッチング結果_YYYYMMDD.html」を自動生成:
  python build_matching_html.py --in マッチング.xlsx
引数なしなら、cwd と ./output から④マッチングDBを持つ最新xlsxを自動検出:
  python build_matching_html.py
明示指定も可:
  python build_matching_html.py --in マッチング.xlsx --out 結果.html [--title "..."] [--subtitle "..."]

シートは名前ではなく先頭セルの内容で特定するため、①〜⑤の採番差異に依存しない。
列はヘッダー名で引くため、列順の差異にも耐える。
"""
import argparse, os, sys, re, glob, html as _html
from datetime import date
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))

# 協業タイプ → 帯色
COOP_COLOR = {
    "共同開発": "#2E5B9A",
    "調達・ツール導入": "#1F8A70",
    "顧客候補": "#C77D0A",
    "技術提供・ライセンス": "#6D5AE0",
    "出資": "#C0392B",
}
# ランク → バッジ枠・スコアバー色
RANK_COLOR = {"A": "#1F8A70", "B": "#2E5B9A", "C": "#C77D0A"}
# 9軸(表示ラベル → Excelヘッダー名)
AXES = [
    ("アセット適合性", "アセット適合性スコア"),
    ("技術実現性", "技術実現性スコア"),
    ("差別化", "差別化スコア"),
    ("業界優位性", "業界優位性スコア"),
    ("業界課題フィット", "業界課題フィットスコア"),
    ("新規性", "新規性スコア"),
    ("ミッション整合性", "ミッション整合性スコア"),
    ("市場規模", "市場規模スコア"),
    ("成長性(CAGR)", "CAGRスコア"),
]


def esc(v):
    if v is None:
        return ""
    return _html.escape(str(v).strip(), quote=False)


def find_sheet(wb, keyword):
    """先頭セル(A1)にkeywordを含むシートを返す。"""
    for ws in wb.worksheets:
        a1 = ws.cell(row=1, column=1).value
        if a1 and keyword in str(a1):
            return ws
    return None


def has_matching_sheet(path):
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ok = find_sheet(wb, "マッチングDB") is not None
        wb.close()
        return ok
    except Exception:
        return False


def autofind_input():
    """cwd と ./output から④マッチングDBを持つ最新の.xlsxを探す。"""
    cands = []
    for pat in ("*.xlsx", os.path.join("output", "*.xlsx")):
        cands += glob.glob(pat)
    cands = [c for c in cands if not os.path.basename(c).startswith("~$")]
    cands.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    for c in cands:
        if has_matching_sheet(c):
            return c
    return None


def default_output(inp):
    """入力と同じディレクトリに「マッチング結果_YYYYMMDD.html」。"""
    d = os.path.dirname(os.path.abspath(inp))
    return os.path.join(d, f"マッチング結果_{date.today():%Y%m%d}.html")


def header_row_index(rows, must_have):
    for i, row in enumerate(rows[:8]):
        vals = [str(c) for c in row if c is not None]
        if all(any(m == v for v in vals) for m in must_have):
            return i
    raise SystemExit(f"header row not found (need {must_have})")


def parse_matches(ws):
    rows = list(ws.iter_rows(values_only=True))
    hi = header_row_index(rows, ["企業名", "No", "タイトル"])
    hdr = rows[hi]
    idx = {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}

    def g(row, name):
        i = idx.get(name)
        return row[i] if (i is not None and i < len(row)) else None

    out = []
    for row in rows[hi + 1:]:
        if not any(c is not None for c in row):
            continue
        if g(row, "企業名") is None and g(row, "タイトル") is None:
            continue
        out.append({k: g(row, k) for k in idx})
    return out


def parse_startup_sites(ws):
    """startup_id → 公式URL"""
    rows = list(ws.iter_rows(values_only=True))
    hi = header_row_index(rows, ["スタートアップID", "企業名"])
    hdr = rows[hi]
    idx = {str(h).strip(): i for i, h in enumerate(hdr) if h is not None}
    sid = idx.get("スタートアップID")
    web = idx.get("Webサイト") or idx.get("公式URL") or idx.get("website")
    m = {}
    if sid is None or web is None:
        return m
    for row in rows[hi + 1:]:
        if sid < len(row) and row[sid]:
            m[str(row[sid]).strip()] = row[web] if web < len(row) else None
    return m


def clean_want(v):
    if not v:
        return ""
    parts = re.split(r"\s*/\s*", str(v))
    keep = [p.strip() for p in parts if p.strip() and "記載ください" not in p and "貴社が展開" not in p]
    return "／".join(keep) if keep else str(v).strip()


def parse_assets(ws):
    """⑤アセット情報 → 出現順のアセットリスト。"""
    rows = list(ws.iter_rows(values_only=True))
    assets = []
    cur = None
    in_func = False
    for row in rows:
        c0 = row[0] if len(row) > 0 else None
        if c0 is None:
            continue
        s = str(c0).strip()
        if s.startswith("■"):
            cur = {"name": s.lstrip("■").strip(), "maturity": "", "gaiyou": "",
                   "specs": [], "funcs": [], "want": ""}
            assets.append(cur)
            in_func = False
            continue
        if cur is None:
            continue
        c1 = row[1] if len(row) > 1 else None
        c2 = row[2] if len(row) > 2 else None
        c4 = row[4] if len(row) > 4 else None
        if s.startswith("成熟度"):
            cur["maturity"] = esc(c1); in_func = False
        elif s.startswith("技術概要"):
            cur["gaiyou"] = esc(c1); in_func = False
        elif s.startswith("[基本]") or s.startswith("[重要]"):
            cur["specs"].append((re.sub(r"^\[(基本|重要)\]\s*", "", s), esc(c1))); in_func = False
        elif s.startswith("スペック項目"):
            in_func = False
        elif s.startswith("機能"):
            in_func = True
        elif s.startswith("拡張可能性") or "WANT" in s:
            cur["want"] = clean_want(c1); in_func = False
        elif s.startswith("NG条件"):
            in_func = False
        elif in_func:
            if c2 or c4:
                cur["funcs"].append({"name": s, "value": esc(c2), "diff": esc(c4)})
    return assets


def asset_overview(a):
    """アセット概要パラグラフ(技術概要 + 提供価値)。"""
    txt = a["gaiyou"]
    vals = [f["value"] for f in a["funcs"] if f["value"]]
    if vals:
        txt = (txt + " 【提供価値】" + "／".join(vals)).strip()
    return txt


def asset_spec_rows(a):
    rows = []
    if a["maturity"]:
        rows.append(("成熟度", a["maturity"]))
    for k, v in a["specs"]:
        if v:
            rows.append((k, v))
    fnames = [f["name"] for f in a["funcs"] if f["name"]]
    if fnames:
        rows.append(("主要機能", "／".join(fnames)))
    return rows


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def split_core_tech(v):
    """使用するコア技術 → (アセット側, スタートアップ側)。"""
    if not v:
        return "", ""
    t = str(v)
    m = re.search(r"/?\s*スタートアップ側[:：]", t)
    if m:
        a = t[:m.start()]
        s = t[m.end():]
        a = re.sub(r"^\s*/?\s*アセット側[:：]\s*", "", a).strip(" /")
        return a.strip(), s.strip()
    return t.strip(), ""


def render(assets, matches, sites, title, subtitle):
    css = open(os.path.join(HERE, "matching_html_style.css"), encoding="utf-8").read()
    js = open(os.path.join(HERE, "matching_html_app.js"), encoding="utf-8").read()

    name2idx = {a["name"]: i for i, a in enumerate(assets)}

    def asset_index(nm):
        nm = (nm or "").strip()
        if nm in name2idx:
            return name2idx[nm]
        for k, i in name2idx.items():
            if nm and (nm in k or k in nm):
                return i
        return None

    # 上部: アセット一覧
    acards = []
    for i, a in enumerate(assets):
        stage = f'<span class="stage">{a["maturity"]}</span>' if a["maturity"] else ""
        want = f'<div class="want"><b>WANT:</b> {a["want"]}</div>' if a["want"] else ""
        acards.append(f'''    <div class="acard" data-f="{i}" title="クリックでこのアセットに絞り込み">
      <h2>{esc(a["name"])}{stage}</h2>
      <p>{asset_overview(a)}</p>
      {want}
    </div>''')

    # フィルタボタン
    counts = {}
    for m in matches:
        ai = asset_index(m.get("アセットID(名)"))
        counts[ai] = counts.get(ai, 0) + 1
    fbtns = [f'<button class="fbtn active" data-f="all">すべて ({len(matches)})</button>']
    for i, a in enumerate(assets):
        fbtns.append(f'<button class="fbtn" data-f="{i}">{esc(a["name"])} ({counts.get(i,0)})</button>')

    # 総合スコア降順
    def score_of(m):
        return num(m.get("総合評価スコア")) or 0.0
    matches_sorted = sorted(matches, key=score_of, reverse=True)

    cards = []
    for m in matches_sorted:
        ai = asset_index(m.get("アセットID(名)"))
        da = str(ai) if ai is not None else "all"
        a = assets[ai] if ai is not None else None

        total = num(m.get("総合評価スコア"))
        rank = (str(m.get("自己評価ランク")).strip() if m.get("自己評価ランク")
                else ("A" if total and total >= 8 else "B" if total and total >= 6 else "C"))
        rc = RANK_COLOR.get(rank, "#2E5B9A")
        coop = esc(m.get("協業タイプ"))
        cc = COOP_COLOR.get(m.get("協業タイプ"), "#5b6577")

        # スタートアップ情報
        sid = str(m.get("スタートアップID") or "").strip()
        site = sites.get(sid)
        prod = esc(m.get("主要プロダクト"))
        prod_sum = esc(m.get("プロダクト概要"))
        prod_disp = prod + (" — " + prod_sum if prod and prod_sum else (prod_sum or ""))
        su_rows = [("事業概要", esc(m.get("事業概要"))),
                   ("主要プロダクト", prod_disp),
                   ("コア技術", esc(m.get("コア技術")))]
        if site:
            su_rows.append(("公式サイト", f'<a href="{esc(site)}" target="_blank" rel="noopener">{esc(site)}</a>'))
        su_html = "\n          ".join(
            f'<div class="row"><span class="k">{k}</span><span class="v">{v}</span></div>'
            for k, v in su_rows if v)

        # 本協業で使う技術
        a_core, s_core = split_core_tech(m.get("使用するコア技術"))
        usetech_a = (f'<div class="usetech usetech-a"><span class="ul">本協業で使う技術</span>{esc(a_core)}</div>'
                     if a_core else "")
        usetech_s = (f'<div class="usetech usetech-s"><span class="ul">本協業で使う技術</span>{esc(s_core)}</div>'
                     if s_core else "")

        # アセット列
        if a:
            spec_html = "".join(
                f'<div class="srow"><span class="sk">{esc(k)}</span><span class="sv">{v}</span></div>'
                for k, v in asset_spec_rows(a))
            asset_col = f'''<div class="tc tc-asset">
          <span class="tc-tag">技術アセット</span>
          <strong class="tc-name">{esc(a["name"])}</strong>
          <p class="tc-sum">{asset_overview(a)}</p>
          <div class="tc-spec"><span class="st">主要スペック</span>{spec_html}</div>
          {usetech_a}
        </div>'''
        else:
            asset_col = f'''<div class="tc tc-asset">
          <span class="tc-tag">技術アセット</span>
          <strong class="tc-name">{esc(m.get("アセットID(名)"))}</strong>
          {usetech_a}
        </div>'''

        # タグ
        tags = [t.strip() for t in re.split(r"[;；]", str(m.get("タグ") or "")) if t.strip()]
        tags_html = "".join(f'<span class="tag">{esc(t)}</span>' for t in tags)

        # 市場規模(一覧)
        mkt = esc(m.get("市場規模(一覧)"))
        cagr = esc(m.get("CAGR(一覧)"))
        mkt_disp = "　/　".join(x for x in [mkt, ("CAGR " + cagr) if cagr else ""] if x)

        # スコアバー
        axes_html = []
        for label, col in AXES:
            v = num(m.get(col))
            if v is None:
                axes_html.append(f'<div class="axis"><span class="axlabel">{label}</span><span class="axna">—</span></div>')
            else:
                axes_html.append(
                    f'<div class="axis"><span class="axlabel">{label}</span>'
                    f'<span class="bar"><span class="fill" style="width:{v*10:.1f}%;background:{rc}"></span></span>'
                    f'<span class="axval">{int(v) if v==int(v) else v}</span></div>')
        scores_html = '<div class="scores">' + "".join(axes_html) + '</div>'

        total_disp = f"{total:.1f}" if total is not None else "—"

        cards.append(f'''    <article class="card" data-asset="{da}" data-score="{total_disp}">
      <div class="card-head">
        <div class="ch-left">
          <span class="coop" style="background:{cc}">{coop}</span>
          <h3>{esc(m.get("タイトル"))}</h3>
        </div>
        <div class="score-badge" style="border-color:{rc}">
          <span class="rank" style="color:{rc}">{esc(rank)}</span>
          <span class="total">{total_disp}</span>
          <span class="totlabel">/10</span>
        </div>
      </div>

      <div class="topcols">
        {asset_col}
        <div class="tc tc-su">
          <span class="tc-tag">スタートアップ</span>
          <strong class="tc-name">{esc(m.get("企業名"))}</strong>
          <div class="su-info">
          {su_html}
          </div>
          {usetech_s}
        </div>
      </div>

      <div class="ideabox">
      <div class="idea-head">協業アイデア</div>
      <div class="tags">{tags_html}</div>
      <div class="idea-body">
        <div class="pt pt-issue"><span class="plbl">課題</span>{esc(m.get("課題"))}</div>
        <div class="pt pt-sol"><span class="plbl">解決策</span>{esc(m.get("解決方法"))}</div>
        <div class="pt pt-strength"><span class="plbl">アイデアの強み</span>{esc(m.get("アイデアの強み"))}</div>
        <div class="idea-mkt"><span class="imk">市場規模</span>{mkt_disp}</div>
      </div>
    </div>

    <details>
      <summary>詳細スコア・想定顧客・収益モデル・評価根拠・リスク等を見る</summary>
      <div class="detail">
          <div class="field scorelbl"><span class="lbl">詳細スコア（0–10・9軸評価）</span></div>
          {scores_html}
          <div class="field"><span class="lbl">想定顧客</span><span class="val">{esc(m.get("想定顧客"))}</span></div>
          <div class="field"><span class="lbl">収益モデル</span><span class="val">{esc(m.get("収益モデル"))}</span></div>
          <div class="field"><span class="lbl">評価根拠</span><span class="val">{esc(m.get("エージェントの思考根拠"))}</span></div>
          <hr>
          <div class="field"><span class="lbl">採用障壁</span><span class="val">{esc(m.get("未使用理由・採用障壁"))}</span></div>
          <div class="field"><span class="lbl">改善点</span><span class="val">{esc(m.get("コア機能の改善点"))}</span></div>
          <div class="field"><span class="lbl">主要リスク</span><span class="val">{esc(m.get("主要リスク"))}</span></div>
          <hr>
          <div class="field"><span class="lbl">市場規模(詳細)</span><span class="val">{esc(m.get("市場規模詳細"))}</span></div>
          <div class="field"><span class="lbl">CAGR(詳細)</span><span class="val">{esc(m.get("CAGR詳細"))}</span></div>
        </div>
      </details>
    </article>''')

    doc = f'''<!doctype html>
<html lang="ja"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{esc(title)}</title>
<style>
{css}</style></head>
<body>
<header class="top">
  <h1>{esc(title)}</h1>
  <p>{esc(subtitle)}</p>
</header>
<div class="wrap">
  <div class="sec">対象技術アセット（{len(assets)}件）<span class="assets-hint">▶ クリックで絞り込み</span></div>
  <section class="assets">
{chr(10).join(acards)}
  </section>
  <div class="sec">マッチング結果（{len(matches)}件）</div>
  <div class="toolbar">
    {''.join(fbtns)}
    <span class="sortsel">スコア降順で表示</span>
  </div>
  <div id="cards">
{chr(10).join(cards)}
  </div>
  <footer>各社の事実欄は公開情報・スタートアップDBに基づく。スコアは評価基準に基づく参考値。</footer>
</div>
<script>
{js}</script>
</body></html>'''
    return doc


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", default=None, help="マッチングExcel(.xlsx)。省略時は自動検出")
    ap.add_argument("--out", dest="out", default=None, help="出力HTML。省略時は入力と同じ場所に「マッチング結果_YYYYMMDD.html」")
    ap.add_argument("--title", default="技術アセット × スタートアップ 協業マッチング結果")
    ap.add_argument("--subtitle", default=None)
    args = ap.parse_args()

    inp = args.inp or autofind_input()
    if not inp:
        sys.exit("マッチングExcelが見つかりません。--in で指定してください(cwd/output に④マッチングDBを持つxlsxが必要)")
    if not os.path.exists(inp):
        sys.exit(f"入力が存在しません: {inp}")
    if not args.inp:
        print(f"[auto] 入力を自動検出: {inp}")
    args.inp = inp
    if not args.out:
        args.out = default_output(inp)

    wb = openpyxl.load_workbook(args.inp, read_only=True, data_only=True)
    ws_match = find_sheet(wb, "マッチングDB")
    ws_asset = find_sheet(wb, "アセット情報")
    ws_su = find_sheet(wb, "スタートアップDB")
    if ws_match is None or ws_asset is None:
        sys.exit("マッチングDB / アセット情報 シートが見つかりません")

    matches = parse_matches(ws_match)
    assets = parse_assets(ws_asset)
    sites = parse_startup_sites(ws_su) if ws_su else {}

    n_su = len(sites) if sites else "?"
    subtitle = args.subtitle or (
        f"技術アセット{len(assets)}件 × スタートアップ{n_su}社 ／ マッチング{len(matches)}件 ／ "
        f"スコアは0–10・9軸評価の平均（A≥8.0 / B≥6.0 / C<6.0）")

    doc = render(assets, matches, sites, args.title, subtitle)
    with open(args.out, "w", encoding="utf-8") as f:
        f.write(doc)
    print(f"OK: {args.out} / assets={len(assets)} matches={len(matches)} startups={n_su} ({len(doc)} chars)")


if __name__ == "__main__":
    main()
