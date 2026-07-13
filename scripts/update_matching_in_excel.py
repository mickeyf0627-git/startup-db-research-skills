# -*- coding: utf-8 -*-
"""update_matching_in_excel.py — 既存のマッチング用Excelに④マッチングDBを正規フォーマットで書き込む

用途: スタートアップDBがJSONLでなく「マッチング用xlsx(③⑤入り)」しか無い場合に、
マッチングJSONLだけを渡して④シートを再生成する。③からのDB転記4列・No・ランク・
総合スコア(一覧)は自動導出。A1マーカー「マッチングDB(N件)」・ヘッダー行(4行目)・
列名/列順は build_startup_db.py の M_COLS をそのまま使うため、能力3(HTML化)と常に互換。

**④シートをopenpyxl等の手書きコードで直接編集してはならない。必ず本スクリプトか
build_startup_db.py を使う**(A1マーカーが消えるとHTML化がシートを発見できなくなる)。

使い方:
  python update_matching_in_excel.py --in <matching.jsonl> --xlsx <既存のマッチング用.xlsx>
  # 検証のみ(書き込まない): --check-only

JSONLフィールド名は SKILL.md「マッチングJSONLの正規スキーマ」に従う(build_startup_db.pyと同一)。
よくある別名(company_name等)はALIASESで自動補正し、その旨を警告表示する。
"""
import argparse, json, os, sys
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from build_startup_db import M_COLS, SCORE_FIELDS, rank_of, hdr, body, FONT  # noqa: E402

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

# 正規フィールド名(M_COLS準拠)。自動導出列(_*)はJSONLに書かない
CANONICAL = {f for _, _, f in M_COLS if f and not f.startswith("_")}
# 過去の実行で観測された別名 → 正規名(後方互換。新規生成では正規名を使うこと)
# ※正規キーが同時に存在する場合は正規キーが優先(別名は欠損の穴埋めのみ)
ALIASES = {
    "company_name": "startup_name", "name": "startup_name",
    "cooperation_type": "coop_type",
    "concept": "summary", "market_size": "market_size_list", "cagr": "cagr_list",
    "total_score": "score_total", "pain_point": "pain", "strength": "idea_strength",
    "score_industry_leadership": "score_industry_advantage",
    "score_industry_fit": "score_industry_pain_fit",
    "score_mission_alignment": "score_mission_fit",
    "unused_reason": "unused_reason_barriers", "improvement": "core_improvement",
    "rationale": "agent_rationale", "risks": "key_risks",
}
# 自動導出・DB転記に相当するため黙って無視するキー
IGNORE = {"no", "rank", "business_overview", "main_product", "product_overview",
          "core_technology", "score_detail"}
SUMMARY_FIELDS = [f for _, _, f in M_COLS if f and f.startswith(("sum_", "score_sum_"))]


def clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "null") else v


def normalize(rec, warns):
    out, alias_vals = {}, {}
    for k, v in rec.items():
        v = clean(v)
        if k in CANONICAL:
            out[k] = v
        elif k == "asset_name":  # asset_idの解決候補として保持(⑤の名称と突合して使う)
            out["_asset_name"] = v
        elif k in ALIASES:
            alias_vals[ALIASES[k]] = v
            warns.add(f"別名フィールド '{k}' → '{ALIASES[k]}' に補正(次回から正規名を使うこと)")
        elif k in IGNORE:
            continue
        else:
            warns.add(f"未知のフィールド '{k}' は無視(SKILL.mdの正規スキーマを確認)")
    for k, v in alias_vals.items():  # 別名は欠損の穴埋めのみ(正規キー優先)
        if not out.get(k):
            out[k] = v
    # スコアは数値化(できなければ空にして警告)
    for f in SCORE_FIELDS:
        if f.startswith("_") or f not in out or out[f] == "":
            continue
        try:
            out[f] = float(out[f])
        except (TypeError, ValueError):
            warns.add(f"スコア '{f}' が数値でない値 '{out[f]}' → 空欄化")
            out[f] = ""
    return out


def read_asset_names(wb):
    """⑤アセット情報の「■ 名称」行からアセット名リストを取得"""
    ws = find_sheet(wb, "アセット情報")
    if ws is None:
        return []
    names = []
    for row in ws.iter_rows(values_only=True):
        c0 = row[0] if row else None
        if c0 and str(c0).strip().startswith("■"):
            names.append(str(c0).strip().lstrip("■").strip())
    return names


def match_asset(nm, names):
    nm = (nm or "").strip()
    if not nm:
        return None
    if nm in names:
        return nm
    for n in names:  # HTML側と同じ部分一致救済
        if nm in n or n in nm:
            return n
    return None


def find_sheet(wb, keyword):
    for ws in wb.worksheets:
        a1 = ws.cell(1, 1).value
        if a1 and keyword in str(a1):
            return ws
    for ws in wb.worksheets:  # A1マーカーが壊れている場合はシート名で救済
        if keyword in ws.title:
            return ws
    return None


def read_db_records(ws):
    """③スタートアップDB(ヘッダー4行目)→ startup_id/企業名 → 転記用dict"""
    rows = list(ws.iter_rows(values_only=True))
    hi = None
    for i, row in enumerate(rows[:8]):
        vals = [str(c) for c in row if c is not None]
        if "スタートアップID" in vals and "企業名" in vals:
            hi = i
            break
    if hi is None:
        raise SystemExit("③スタートアップDBのヘッダー行が見つかりません")
    idx = {str(h).strip(): i for i, h in enumerate(rows[hi]) if h is not None}
    need = {"business_summary": "事業概要", "product_name": "主要プロダクト",
            "product_summary": "プロダクト概要", "core_tech_description": "コア技術"}
    by_key = {}
    for row in rows[hi + 1:]:
        d = {}
        for field, header in need.items():
            j = idx.get(header)
            d[field] = row[j] if (j is not None and j < len(row)) else None
        for keyh in ("スタートアップID", "企業名", "企業名(日)"):
            j = idx.get(keyh)
            k = row[j] if (j is not None and j < len(row)) else None
            if k:
                by_key[str(k).strip().lower()] = d
    return by_key


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True, help="マッチングJSONL")
    ap.add_argument("--xlsx", required=True, help="③⑤入りの既存マッチング用xlsx(④を上書き再生成)")
    ap.add_argument("--check-only", action="store_true", help="検証のみで書き込まない")
    a = ap.parse_args()

    warns = set()
    recs = []
    with open(a.inp, encoding="utf-8") as f:
        for line in f:
            if line.strip():
                recs.append(normalize(json.loads(line), warns))
    if not recs:
        raise SystemExit("JSONLが空です")

    missing_sum = sum(1 for r in recs if not any(r.get(f) for f in SUMMARY_FIELDS))
    if missing_sum:
        warns.add(f"[品質] 要約ブロック13列が空の行が{missing_sum}件(SKILL.md: 要約はエージェントが本編から凝縮して記入)")
    for i, r in enumerate(recs, 1):
        for req in ("startup_id", "title", "summary", "score_total", "agent_rationale"):
            if not r.get(req):
                warns.add(f"行{i}: 必須フィールド '{req}' が空")

    # スコアの自動導出: score_total=9軸平均 / score_sum_*4種=本編スコアの転記。
    # モデルに算術をさせない(記入値があっても計算値で上書き)。JSONLでは省略してよい
    MAIN_SCORES = [f for f in SCORE_FIELDS
                   if not f.startswith(("_", "score_sum_")) and f != "score_total"]
    SUM_SCORE_MAP = {"score_sum_uniqueness": "score_novelty",
                     "score_sum_mission": "score_mission_fit",
                     "score_sum_market_size": "score_market_size",
                     "score_sum_growth": "score_cagr"}
    n_drv = 0
    for r in recs:
        vals = [r.get(f) for f in MAIN_SCORES]
        if all(isinstance(v, float) for v in vals):
            total = round(sum(vals) / len(vals), 1)
            if r.get("score_total") != total:
                n_drv += 1
            r["score_total"] = total
        for sf, mf in SUM_SCORE_MAP.items():
            if isinstance(r.get(mf), float):
                r[sf] = r[mf]
    if n_drv:
        warns.add(f"score_totalを9軸平均で再計算し{n_drv}件を上書き(score_sum_*4種も本編スコアから自動転記)")

    # [品質]警告(内容レベル): 書き込みは止めないが、SKILL.md「品質改善パス」の対象行特定に使う。
    # 基準はSKILL.md「出力品質基準」(良品サンプルの実測レンジ)に対応
    import re
    n_abc = n_smry = n_ph = n_ban = n_len = n_title = n_risk = n_int = n_slug = 0
    score_patterns, score_shapes, ranks_seen = set(), set(), set()
    # 「後で埋める」つもりの定型プレースホルダー。空欄(=事実性の鉄則で正)とは別に検出する
    PLACEHOLDER_RE = re.compile(r"(詳細分析|後で|追って|TBD|todo|要検討|検討中|分析中|予定|要記入|未記入)", re.I)
    PLACEHOLDER_FIELDS = ("core_improvement", "key_risks", "unused_reason_barriers",
                          "market_size_detail", "cagr_detail")
    # 禁止語(定型文の兆候)と、良品サンプル実測レンジに基づく文字数フロア
    BANNED = ("相乗効果", "技術融合", "唯一の実現", "による課題解決", "協業による事業展開")
    BAN_FIELDS = ("summary", "solution", "idea_strength")
    LEN_FLOORS = {"summary": 120, "pain": 70, "solution": 80,
                  "idea_strength": 50, "agent_rationale": 150}
    THINK_FIELDS = ("pain", "solution", "idea_strength", "summary", "agent_rationale")
    # スラッグ漏れ検出で除外する正規の軸名・フィールド名(思考根拠での言及は正当)
    LEGIT_TOKENS = sorted(set(list(SCORE_FIELDS) + list(CANONICAL) +
                              ["asset_fit", "pain_fit", "tech_feasibility", "mission_fit",
                               "industry_advantage", "industry_pain_fit", "score_total"]),
                          key=len, reverse=True)
    LEGIT_RE = re.compile("|".join(re.escape(t) for t in LEGIT_TOKENS if t))
    dup_seen = {}  # (field, 正規化文) -> 出現行数。企業名を差し替えただけのコピペを検出
    for r in recs:
        rat = str(r.get("agent_rationale") or "")
        if rat and not re.search(r"[(（][abc][)）]", rat):
            n_abc += 1  # 原則2: どの判定式を満たすかの明記が必須
        smry = str(r.get("summary") or "")
        if smry and (len(smry) < 30 or "協業による事業展開" in smry):
            n_smry += 1
        if any(PLACEHOLDER_RE.search(str(r.get(f) or "")) for f in PLACEHOLDER_FIELDS):
            n_ph += 1
        if any(b in str(r.get(f) or "") for f in BAN_FIELDS for b in BANNED):
            n_ban += 1
        if any(r.get(f) and len(str(r.get(f))) < floor for f, floor in LEN_FLOORS.items()):
            n_len += 1
        if "×" in str(r.get("title") or ""):
            n_title += 1  # タイトルは成果物を名指す(「X × Y」機械結合は不可)
        kr = str(r.get("key_risks") or "")
        if kr and ("①" not in kr or "②" not in kr):
            n_risk += 1  # 主要リスクは①②③の番号列挙で3〜4件
        svals = [r.get(f) for f in MAIN_SCORES if isinstance(r.get(f), float)]
        if any(v != int(v) for v in svals):
            n_int += 1  # 9軸スコアは0-10の整数
        if any(re.search(r"[a-z]+_[a-z]+", LEGIT_RE.sub("", str(r.get(f) or "")))
               for f in THINK_FIELDS):
            n_slug += 1  # 英語スラッグ(some_asset_slug等)の本文への漏れ(正規の軸名言及は除外)
        # 行間コピペ検出: 企業名・アセット名を除去した思考フィールドが他行と一致したら定型量産
        names = [str(r.get("startup_name") or ""), str(r.get("asset_id") or "")]
        for f in THINK_FIELDS:
            t = str(r.get(f) or "")
            for nm in names:
                if nm:
                    t = t.replace(nm, "")
            t = re.sub(r"\s+", "", t)
            if len(t) >= 20:
                dup_seen[(f, t)] = dup_seen.get((f, t), 0) + 1
        if svals:
            score_patterns.add(tuple(svals))
            m = sum(svals) / len(svals)
            score_shapes.add(tuple(round(v - m, 1) for v in svals))  # 平均を引いた「形」
        t = r.get("score_total")
        ranks_seen.add(rank_of(t if isinstance(t, float) else None))
    n_dup = sum(c for c in dup_seen.values() if c >= 2)
    if n_abc:
        warns.add(f"[品質] 思考根拠に原則2のa/b/c判定の明記がない行が{n_abc}件")
    if n_smry:
        warns.add(f"[品質] 概要が定型文または30字未満の行が{n_smry}件")
    if n_ph:
        warns.add(f"[品質] 採用障壁/改善点/主要リスク/市場規模詳細等に定型プレースホルダー"
                  f"(『詳細分析』『予定』等)が残る行が{n_ph}件(中身を書くか、事実不明なら空欄に)")
    if n_ban:
        warns.add(f"[品質] 概要/解決方法/強みに禁止語(相乗効果・技術融合・唯一の実現・による課題解決等)を含む行が{n_ban}件")
    if n_len:
        warns.add(f"[品質] 文字数が品質基準レンジ未満の行が{n_len}件(概要<120/課題<70/解決<80/強み<50/根拠<150字)")
    if n_title:
        warns.add(f"[品質] タイトルが「X × Y」機械結合の行が{n_title}件(協業の成果物を名指すこと)")
    if n_risk:
        warns.add(f"[品質] 主要リスクが①②の番号列挙になっていない行が{n_risk}件(3〜4件を列挙)")
    if n_int:
        warns.add(f"[品質] 9軸スコアに非整数を含む行が{n_int}件(0-10の整数で採点)")
    if n_slug:
        warns.add(f"[品質] 本文に英語スラッグ(snake_case)が漏れている行が{n_slug}件")
    if n_dup:
        warns.add(f"[品質] 企業名を差し替えただけの行間コピペが{n_dup}箇所(課題/解決/強み/概要/根拠は行ごとに固有の内容)")
    if len(recs) >= 3 and len(score_patterns) == 1:
        warns.add(f"[品質] 全{len(recs)}行のスコアが同一パターン(個別評価が行われていない疑い)")
    elif len(recs) >= 5 and len(score_shapes) == 1:
        warns.add(f"[品質] 全{len(recs)}行のスコアが同一の形(定数オフセットの等差列。個別評価が行われていない疑い)")
    if len(recs) >= 5 and len(ranks_seen - {None}) == 1:
        warns.add(f"[品質] 全{len(recs)}行が同一ランク(正直採点ならランクは分布する。無理にBに乗せない)")

    wb = openpyxl.load_workbook(a.xlsx)
    db_ws = find_sheet(wb, "スタートアップDB")
    if db_ws is None:
        raise SystemExit("③スタートアップDBシートが見つかりません")
    by_key = read_db_records(db_ws)
    linked = sum(1 for r in recs if str(r.get("startup_id", "")).strip().lower() in by_key)
    if linked < len(recs):
        warns.add(f"③と結合できないstartup_idが{len(recs)-linked}件(転記4列が空になる)")

    # アセットID(名)を⑤の名称に解決(HTMLの絞り込みはこの一致に依存する)
    anames = read_asset_names(wb)
    if anames:
        for i, r in enumerate(recs, 1):
            resolved = match_asset(r.get("asset_id"), anames) or match_asset(r.get("_asset_name"), anames)
            if resolved is None:
                warns.add(f"行{i}: asset_id '{r.get('asset_id')}' が⑤のどのアセット名とも一致しない(HTML絞り込み対象外になる)")
            elif resolved != r.get("asset_id"):
                warns.add(f"asset_id '{r.get('asset_id')}' → ⑤の名称 '{resolved}' に解決")
                r["asset_id"] = resolved
            r.pop("_asset_name", None)
    else:
        warns.add("⑤アセット情報シートが見つからずアセット名の整合チェックをスキップ")
        for r in recs:
            r.pop("_asset_name", None)

    print(f"検証: {len(recs)}件 / ③との結合 {linked}/{len(recs)} / ⑤アセット名 {len(anames)}件")
    for w in sorted(warns):
        print(f"  [警告] {w}")
    if a.check_only:
        print("(--check-only のため書き込みなし)")
        return

    # ④シートを同位置に再生成
    ms = find_sheet(wb, "マッチングDB")
    if ms is not None:
        idx_pos, name = wb.worksheets.index(ms), ms.title
        wb.remove(ms)
    else:
        idx_pos, name = len(wb.worksheets), "④マッチングDB"
    ms = wb.create_sheet(name, idx_pos)

    c = ms.cell(1, 1, f"マッチングDB({len(recs)}件)")
    c.font = Font(name=FONT, bold=True, size=14, color="1F3864")
    s = ms.cell(2, 1, "アセット×スタートアップの協業アイデア評価(アイデア評価DB互換フォーマット)。スコアは0-10の数値。ランクは総合評価スコアから自動導出(A>=8.0/B>=6.0/C<6.0)。")
    s.font = Font(name=FONT, size=9, color="595959")

    r = 4
    for i, (h, w, _) in enumerate(M_COLS, 1):
        hdr(ms.cell(r, i, h), fill="7F6000")
        ms.column_dimensions[get_column_letter(i)].width = w
    rank_fill = {"A": "E2EFDA", "B": "DCE6F1", "C": "FCE4D6"}
    for no, d in enumerate(recs, 1):
        r += 1
        total = d.get("score_total") if isinstance(d.get("score_total"), float) else None
        rank = rank_of(total)
        src = by_key.get(str(d.get("startup_id", "")).strip().lower()) or {}
        for i, (h, w, field) in enumerate(M_COLS, 1):
            if field == "_no":
                v = no
            elif field == "_total_list":
                v = total if total is not None else ""
            elif field == "_rank":
                v = rank or ""
            elif field and field.startswith("_db_"):
                v = src.get({"_db_business_summary": "business_summary",
                             "_db_product_name": "product_name",
                             "_db_product_summary": "product_summary",
                             "_db_core_tech": "core_tech_description"}[field]) or ""
            else:
                v = d.get(field, "")
                if isinstance(v, list):
                    v = "; ".join(str(x) for x in v)
            c = ms.cell(r, i, v)
            body(c, center=(field in SCORE_FIELDS or field in ("_no", "_rank")))
            if field == "_rank" and rank:
                c.fill = PatternFill("solid", start_color=rank_fill[rank])
                c.font = Font(name=FONT, size=10, bold=True)
    ms.freeze_panes = "E5"
    ms.auto_filter.ref = f"A4:{get_column_letter(len(M_COLS))}{max(r, 5)}"

    wb.save(a.xlsx)
    ranks = {}
    for d in recs:
        t = d.get("score_total")
        rk = rank_of(t if isinstance(t, float) else None) or "?"
        ranks[rk] = ranks.get(rk, 0) + 1
    print(f"保存: {a.xlsx} (④マッチングDB {len(recs)}件, ランク: "
          + " / ".join(f"{k}={v}" for k, v in sorted(ranks.items())) + ")")


if __name__ == "__main__":
    main()
